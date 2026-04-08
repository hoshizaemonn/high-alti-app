"""Data upload page — handles payroll (Excel/CSV), expense (CSV), and hacomono data."""

import streamlit as st
import pandas as pd
import openpyxl
import io
import csv
import re
from datetime import datetime, date
from database import (
    save_payroll_data, save_expense_data, save_revenue_data,
    save_member_data, save_monthly_summary, save_sales_detail,
    save_budget_data, check_budget_exists, BUDGET_ITEMS,
    save_square_sales,
    classify_sale_category, SALES_CATEGORIES,
    upsert_override, STORES, HQ_STORE, EXPENSE_CATEGORIES,
    save_amazon_orders, get_amazon_order_count,
    detect_store_from_address, match_amazon_breakdown,
    check_payroll_exists, check_expense_exists,
    check_member_exists, check_sales_detail_exists,
    check_monthly_summary_exists, upsert_expense_rule,
)

# Store options including HQ for employee assignment
STORE_OPTIONS_WITH_HQ = STORES + [HQ_STORE]
from store_logic import resolve_store, apply_ratio
from expense_logic import classify_expense


# hacomono store name mapping: full name → short name used in the app
HACOMONO_STORE_MAP = {
    "ハイアルチ東日本橋スタジオ": "東日本橋",
    "ハイアルチ春日スタジオ": "春日",
    "ハイアルチ船橋スタジオ": "船橋",
    "ハイアルチ巣鴨スタジオ": "巣鴨",
    "ハイアルチ祖師ヶ谷大蔵スタジオ": "祖師ヶ谷大蔵",
    "ハイアルチ下北沢スタジオ": "下北沢",
    "ハイアルチ中目黒スタジオ": "中目黒",
    "ハイアルチ東陽町スタジオ": "東陽町",
}


def _map_hacomono_store(full_name: str) -> str:
    """Map hacomono full store name to short name."""
    full_name = full_name.strip()
    if full_name in HACOMONO_STORE_MAP:
        return HACOMONO_STORE_MAP[full_name]
    # Fallback: strip prefix/suffix
    short = full_name.replace("ハイアルチ", "").replace("スタジオ", "").strip()
    return short if short else full_name


def _parse_date_loose(val: str) -> date | None:
    """Parse a date string in various formats. Returns None if unparseable or empty."""
    if not val or not val.strip():
        return None
    val = val.strip()
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _is_in_month(dt: date | None, year: int, month: int) -> bool:
    """Check if a date falls within the given year/month."""
    if dt is None:
        return False
    return dt.year == year and dt.month == month


def _parse_ml001_csv(
    file_bytes: bytes, year: int, month: int, target_store: str
) -> tuple[list[dict], dict]:
    """Parse hacomono ML001 member CSV.

    Returns (records, summary_info).
    CSV is UTF-8-sig with 58 columns. Uses 0-indexed column positions as fallback.
    """
    # Decode
    text = None
    for enc in ["utf-8-sig", "utf-8", "cp932"]:
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSVのエンコーディングを判定できませんでした")

    reader = csv.reader(io.StringIO(text))
    header = next(reader)

    # Build column index map (header name → index)
    hmap = {}
    for i, h in enumerate(header):
        hmap[h.strip()] = i

    # Key column indices — try header name first, fall back to known positions
    def _col(name: str, fallback: int) -> int:
        return hmap.get(name, fallback)

    idx_member_id = _col("メンバーID", 0)
    idx_member_name = _col("氏名", 2)
    idx_trial_date = _col("無料体験会 受講日時", 37)
    idx_first_trial = _col("トライアル 初回受講日時", 38)
    idx_join_date = _col("入会日時", 39)
    idx_member_store = _col("メンバー所属店舗名", 44)
    idx_plan_name = _col("契約プラン名", 47)
    idx_current_store = _col("所属店舗名", 49)
    idx_plan_contract_date = _col("プラン契約日", 50)
    idx_plan_end_date = _col("プラン契約適用終了日", 52)
    idx_initial_plan = _col("初回契約プラン", 55)
    idx_tenure = _col("在籍期間", 56)

    today = date.today()
    records = []
    empty_store_count = 0
    total_count = 0
    active_count = 0
    suspended_count = 0
    new_count = 0
    trial_count = 0
    plan_counts = {}
    active_plan_counts = {}
    personal_ticket_count = 0

    for row in reader:
        if len(row) < 10:
            continue

        # Extract fields safely
        def _get(idx: int) -> str:
            return row[idx].strip() if idx < len(row) else ""

        member_id = _get(idx_member_id)
        member_name = _get(idx_member_name)
        plan_name = _get(idx_plan_name)

        # Skip rows with no plan name (empty rows)
        if not plan_name:
            continue

        # Determine store
        store_full = _get(idx_current_store)
        if not store_full:
            store_full = _get(idx_member_store)

        if store_full:
            store_short = _map_hacomono_store(store_full)
        else:
            # Empty store name — assign to target store (they belong to this store's export)
            store_short = target_store
            empty_store_count += 1

        total_count += 1

        # Parse dates
        trial_dt = _parse_date_loose(_get(idx_trial_date))
        first_trial_dt = _parse_date_loose(_get(idx_first_trial))
        join_dt = _parse_date_loose(_get(idx_join_date))
        plan_contract_dt = _parse_date_loose(_get(idx_plan_contract_date))
        plan_end_dt = _parse_date_loose(_get(idx_plan_end_date))

        tenure = _get(idx_tenure)
        initial_plan = _get(idx_initial_plan)
        join_date_str = _get(idx_join_date)
        trial_date_str = _get(idx_trial_date)
        first_trial_str = _get(idx_first_trial)
        plan_end_str = _get(idx_plan_end_date)

        # Determine is_active: active unless 休会 in plan name, or plan_end_date is in the past
        is_suspended = "休会" in plan_name
        has_ended = plan_end_dt is not None and plan_end_dt < today
        is_active = 0 if (is_suspended or has_ended) else 1

        # Determine is_new: tenure is "1ヶ月目" OR plan_contract_date is in the selected month
        is_new = 0
        if tenure == "1ヶ月目":
            is_new = 1
        elif _is_in_month(plan_contract_dt, year, month):
            is_new = 1

        # Determine had_trial: trial date or first trial date falls within selected month
        had_trial = 0
        if _is_in_month(trial_dt, year, month) or _is_in_month(first_trial_dt, year, month):
            had_trial = 1

        # Track パーソナルチケット
        if "パーソナル" in plan_name and "チケット" in plan_name:
            personal_ticket_count += 1

        records.append({
            "year": year,
            "month": month,
            "store_name": store_short,
            "member_id": member_id,
            "member_name": member_name,
            "plan_name": plan_name,
            "join_date": join_date_str,
            "tenure": tenure,
            "is_active": is_active,
            "is_new": is_new,
            "had_trial": had_trial,
            "plan_end_date": plan_end_str,
            "trial_date": trial_date_str,
            "first_trial_date": first_trial_str,
            "initial_plan": initial_plan,
        })

        plan_counts[plan_name] = plan_counts.get(plan_name, 0) + 1
        if is_active:
            active_count += 1
            active_plan_counts[plan_name] = active_plan_counts.get(plan_name, 0) + 1
        if is_suspended:
            suspended_count += 1
        if is_new:
            new_count += 1
        if had_trial:
            trial_count += 1

    summary_info = {
        "total": total_count,
        "active": active_count,
        "suspended": suspended_count,
        "new": new_count,
        "trial": trial_count,
        "plan_counts": plan_counts,
        "active_plan_counts": active_plan_counts,
        "empty_store_count": empty_store_count,
        "personal_ticket": personal_ticket_count,
        "store": target_store,
    }

    return records, summary_info


def _render_ml001_summary(info: dict):
    """Render summary after ML001 import."""
    st.markdown("---")
    st.markdown("### 取込結果サマリー")

    # KPI row
    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        st.metric("全会員数", f"{info['total']}名")
    with k2:
        st.metric("有効在籍数", f"{info['active']}名")
    with k3:
        st.metric("休会", f"{info['suspended']}名")
    with k4:
        st.metric("新規入会（当月）", f"{info['new']}名")
    with k5:
        st.metric("体験（当月）", f"{info['trial']}名")

    # Plan breakdown table
    st.markdown("---")
    col_plan, col_active_plan = st.columns(2)

    with col_plan:
        st.markdown("**プラン別 全会員数**")
        if info["plan_counts"]:
            plan_data = sorted(info["plan_counts"].items(), key=lambda x: -x[1])
            plan_df = pd.DataFrame(plan_data, columns=["プラン名", "会員数"])
            total = sum(v for _, v in plan_data)
            plan_df["構成比"] = plan_df["会員数"].apply(lambda x: f"{x / total * 100:.1f}%")
            st.dataframe(plan_df, use_container_width=True, hide_index=True)

    with col_active_plan:
        st.markdown("**プラン別 有効在籍数**")
        if info["active_plan_counts"]:
            active_data = sorted(info["active_plan_counts"].items(), key=lambda x: -x[1])
            active_df = pd.DataFrame(active_data, columns=["プラン名", "会員数"])
            active_total = sum(v for _, v in active_data)
            active_df["構成比"] = active_df["会員数"].apply(lambda x: f"{x / active_total * 100:.1f}%")
            st.dataframe(active_df, use_container_width=True, hide_index=True)


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _detect_year_month_from_filename(filename: str) -> tuple[int | None, int | None]:
    """Try to detect year and month from filename patterns like 2026年02月 or 202602."""
    # Pattern: 2026年02月 or 2026年2月
    m = re.search(r'(\d{4})年(\d{1,2})月', filename)
    if m:
        return int(m.group(1)), int(m.group(2))
    # Pattern: 202602
    m = re.search(r'(\d{4})(\d{2})', filename)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 2020 <= y <= 2030 and 1 <= mo <= 12:
            return y, mo
    return None, None


def _parse_payroll_from_worksheet(ws, year: int, month: int) -> tuple[list[dict], list[dict]]:
    """Parse a single worksheet (payroll data) into records."""
    records = []
    unresolved = []

    for row_idx in range(2, ws.max_row + 1):
        emp_id_raw = ws.cell(row=row_idx, column=1).value
        if emp_id_raw is None:
            continue

        emp_id_str = str(emp_id_raw).strip()

        # Skip header-like rows
        if emp_id_str.startswith("【"):
            continue

        emp_name = ws.cell(row=row_idx, column=2).value or ""
        if emp_name == "-":
            continue

        contract_type = ws.cell(row=row_idx, column=6).value or ""

        # Parse all numeric fields
        work_days_weekday = _safe_float(ws.cell(row=row_idx, column=7).value)
        work_days_holiday = _safe_float(ws.cell(row=row_idx, column=8).value)
        work_days_legal = _safe_float(ws.cell(row=row_idx, column=9).value)
        scheduled_hours = _safe_float(ws.cell(row=row_idx, column=13).value)

        overtime_hours = sum(
            _safe_float(ws.cell(row=row_idx, column=c).value)
            for c in range(16, 25)
        )

        base_salary = _safe_float(ws.cell(row=row_idx, column=28).value)
        position_allowance = _safe_float(ws.cell(row=row_idx, column=29).value)
        overtime_pay = _safe_float(ws.cell(row=row_idx, column=33).value)
        commute_taxable = _safe_float(ws.cell(row=row_idx, column=45).value)
        commute_nontax = _safe_float(ws.cell(row=row_idx, column=46).value)
        taxable_total = _safe_float(ws.cell(row=row_idx, column=52).value)
        gross_total = _safe_float(ws.cell(row=row_idx, column=56).value)

        health_ins_co = _safe_float(ws.cell(row=row_idx, column=90).value)
        care_ins_co = _safe_float(ws.cell(row=row_idx, column=91).value)
        pension_co = _safe_float(ws.cell(row=row_idx, column=92).value)
        child_co = _safe_float(ws.cell(row=row_idx, column=93).value)
        pension_fund_co = _safe_float(ws.cell(row=row_idx, column=94).value)
        employment_ins_co = _safe_float(ws.cell(row=row_idx, column=95).value)
        workers_comp_co = _safe_float(ws.cell(row=row_idx, column=96).value)
        general_co = _safe_float(ws.cell(row=row_idx, column=97).value)

        store_assignments = resolve_store(emp_id_str)

        if not store_assignments:
            unresolved.append({
                "employee_id": emp_id_str,
                "employee_name": emp_name,
                "contract_type": contract_type,
                "gross_total": gross_total,
            })
            continue

        for assignment in store_assignments:
            ratio = assignment["ratio"]
            records.append({
                "year": year,
                "month": month,
                "employee_id": emp_id_str,
                "employee_name": emp_name,
                "contract_type": contract_type,
                "store_name": assignment["store_name"],
                "ratio": ratio,
                "work_days_weekday": apply_ratio(work_days_weekday, ratio),
                "work_days_holiday": apply_ratio(work_days_holiday, ratio),
                "work_days_legal_holiday": apply_ratio(work_days_legal, ratio),
                "scheduled_hours": apply_ratio(scheduled_hours, ratio),
                "overtime_hours": apply_ratio(overtime_hours, ratio),
                "base_salary": apply_ratio(base_salary, ratio),
                "position_allowance": apply_ratio(position_allowance, ratio),
                "overtime_pay": apply_ratio(overtime_pay, ratio),
                "commute_taxable": apply_ratio(commute_taxable, ratio),
                "commute_nontax": apply_ratio(commute_nontax, ratio),
                "taxable_total": apply_ratio(taxable_total, ratio),
                "gross_total": apply_ratio(gross_total, ratio),
                "health_insurance_co": apply_ratio(health_ins_co, ratio),
                "care_insurance_co": apply_ratio(care_ins_co, ratio),
                "pension_co": apply_ratio(pension_co, ratio),
                "child_contribution_co": apply_ratio(child_co, ratio),
                "pension_fund_co": apply_ratio(pension_fund_co, ratio),
                "employment_insurance_co": apply_ratio(employment_ins_co, ratio),
                "workers_comp_co": apply_ratio(workers_comp_co, ratio),
                "general_contribution_co": apply_ratio(general_co, ratio),
            })

    return records, unresolved


def parse_payroll_excel(file_bytes: bytes, year: int, month: int) -> tuple[list[dict], list[dict]]:
    """Parse payroll Excel (.xlsx) file."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    detail_sheet = None
    for name in wb.sheetnames:
        if "支給控除一覧表" in name:
            detail_sheet = wb[name]
            break
    if detail_sheet is None:
        detail_sheet = wb[wb.sheetnames[0]]

    return _parse_payroll_from_worksheet(detail_sheet, year, month)


def parse_payroll_csv(file_bytes: bytes, year: int, month: int) -> tuple[list[dict], list[dict]]:
    """Parse payroll CSV file. Try multiple encodings."""
    for enc in ["cp932", "utf-8", "utf-8-sig"]:
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("CSVのエンコーディングを判定できませんでした")

    reader = csv.reader(io.StringIO(text))
    header = next(reader)

    records = []
    unresolved = []

    for row in reader:
        if len(row) < 10:
            continue

        emp_id_str = str(row[0]).strip()
        if not emp_id_str or emp_id_str.startswith("【"):
            continue

        emp_name = row[1].strip() if len(row) > 1 else ""
        if emp_name == "-":
            continue

        contract_type = row[5].strip() if len(row) > 5 else ""

        # Parse numeric columns by index (same as Excel column numbers - 1)
        def col(idx):
            return _safe_float(row[idx]) if len(row) > idx else 0.0

        work_days_weekday = col(6)
        work_days_holiday = col(7)
        work_days_legal = col(8)
        scheduled_hours = col(12)
        overtime_hours = sum(col(c) for c in range(15, 24))

        base_salary = col(27)
        position_allowance = col(28)
        overtime_pay = col(32)
        commute_taxable = col(44)
        commute_nontax = col(45)
        taxable_total = col(51)
        gross_total = col(55)

        health_ins_co = col(89)
        care_ins_co = col(90)
        pension_co = col(91)
        child_co = col(92)
        pension_fund_co = col(93)
        employment_ins_co = col(94)
        workers_comp_co = col(95)
        general_co = col(96)

        store_assignments = resolve_store(emp_id_str)

        if not store_assignments:
            unresolved.append({
                "employee_id": emp_id_str,
                "employee_name": emp_name,
                "contract_type": contract_type,
                "gross_total": gross_total,
            })
            continue

        for assignment in store_assignments:
            ratio = assignment["ratio"]
            records.append({
                "year": year,
                "month": month,
                "employee_id": emp_id_str,
                "employee_name": emp_name,
                "contract_type": contract_type,
                "store_name": assignment["store_name"],
                "ratio": ratio,
                "work_days_weekday": apply_ratio(work_days_weekday, ratio),
                "work_days_holiday": apply_ratio(work_days_holiday, ratio),
                "work_days_legal_holiday": apply_ratio(work_days_legal, ratio),
                "scheduled_hours": apply_ratio(scheduled_hours, ratio),
                "overtime_hours": apply_ratio(overtime_hours, ratio),
                "base_salary": apply_ratio(base_salary, ratio),
                "position_allowance": apply_ratio(position_allowance, ratio),
                "overtime_pay": apply_ratio(overtime_pay, ratio),
                "commute_taxable": apply_ratio(commute_taxable, ratio),
                "commute_nontax": apply_ratio(commute_nontax, ratio),
                "taxable_total": apply_ratio(taxable_total, ratio),
                "gross_total": apply_ratio(gross_total, ratio),
                "health_insurance_co": apply_ratio(health_ins_co, ratio),
                "care_insurance_co": apply_ratio(care_ins_co, ratio),
                "pension_co": apply_ratio(pension_co, ratio),
                "child_contribution_co": apply_ratio(child_co, ratio),
                "pension_fund_co": apply_ratio(pension_fund_co, ratio),
                "employment_insurance_co": apply_ratio(employment_ins_co, ratio),
                "workers_comp_co": apply_ratio(workers_comp_co, ratio),
                "general_contribution_co": apply_ratio(general_co, ratio),
            })

    return records, unresolved


def parse_expense_csv(file_bytes: bytes, encoding: str = "cp932") -> list[dict]:
    """Parse PayPay銀行 CSV and return classified expense records."""
    text = file_bytes.decode(encoding)
    reader = csv.reader(io.StringIO(text))
    header = next(reader)

    # Guard: detect if this is an Amazon CSV instead of PayPay
    header_str = ",".join(header)
    if "注文番号" in header_str or "ASIN" in header_str:
        raise ValueError("これはAmazon注文履歴CSVです。経費CSVではありません。先にAmazon CSVとして取り込んでください。")

    records = []
    for row in reader:
        if len(row) < 12:
            continue

        year = int(row[0])
        month = int(row[1])
        day = int(row[2])
        description = row[7].strip()
        amount_str = row[8].strip()
        deposit_str = row[9].strip()

        amount = float(amount_str) if amount_str else 0.0
        deposit = float(deposit_str) if deposit_str else 0.0

        category, is_revenue = classify_expense(description)

        records.append({
            "year": year,
            "month": month,
            "day": day,
            "description": description,
            "amount": amount,
            "deposit": deposit,
            "category": category if category != "_収入" else "_収入",
            "is_revenue": 1 if is_revenue else 0,
        })

    return records


def _shorten_product_name(name: str, max_len: int = 30) -> str:
    """Shorten an Amazon product name for display."""
    if not name:
        return ""
    # Remove common noise patterns
    short = re.sub(r'\s*[\[【（(].*?[\]】）)]', '', name)
    short = short.strip()
    if len(short) > max_len:
        short = short[:max_len] + "…"
    return short


def parse_amazon_csv(file_bytes: bytes) -> list[dict]:
    """Parse Amazon order history CSV and return order records."""
    text = None
    for enc in ["utf-8-sig", "utf-8", "cp932"]:
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSVのエンコーディングを判定できませんでした")

    reader = csv.reader(io.StringIO(text))
    header = next(reader)

    # Build column index map
    col_map = {}
    for i, h in enumerate(header):
        h_clean = h.strip().strip('\ufeff')
        if h_clean == '注文日':
            col_map['order_date'] = i
        elif h_clean == '注文番号':
            col_map['order_id'] = i
        elif h_clean == '商品名':
            col_map['product_name'] = i
        elif h_clean == '注文状況':
            col_map['status'] = i
        elif h_clean == '注文の合計（税込）':
            col_map['order_total'] = i
        elif h_clean == '支払い金額':
            col_map['payment_amount'] = i
        elif h_clean == '支払い確定日':
            col_map['payment_date'] = i
        elif h_clean == '商品の小計（税込）':
            col_map['item_amount'] = i
        elif h_clean.startswith('発送商品の') and '合計' in h_clean and '税込' in h_clean:
            col_map['ship_total'] = i

    def _parse_int(val_str):
        if not val_str:
            return 0
        cleaned = val_str.strip().replace(',', '').replace('"', '').replace('=', '').replace('￥', '').replace('¥', '')
        if not cleaned or cleaned == '該当なし':
            return 0
        try:
            return int(cleaned)
        except ValueError:
            return 0

    orders = []
    for row in reader:
        if len(row) < 10:
            continue

        # Skip cancelled orders
        status_idx = col_map.get('status', -1)
        if 0 <= status_idx < len(row) and 'キャンセル' in row[status_idx]:
            continue

        order_date = row[col_map.get('order_date', 0)].strip() if 'order_date' in col_map else ""
        order_id = row[col_map.get('order_id', 1)].strip() if 'order_id' in col_map else ""
        product_name = row[col_map.get('product_name', -1)].strip() if 'product_name' in col_map else ""

        # Order total (what PayPay charges)
        order_total = 0
        for key in ['payment_amount', 'order_total']:
            if key in col_map:
                order_total = _parse_int(row[col_map[key]])
                if order_total > 0:
                    break

        # Individual item amount
        item_amount = 0
        for key in ['item_amount', 'ship_total']:
            if key in col_map:
                item_amount = _parse_int(row[col_map[key]])
                if item_amount > 0:
                    break

        # Payment date
        payment_date = row[col_map['payment_date']].strip() if 'payment_date' in col_map and col_map['payment_date'] < len(row) else ""

        # Find delivery address - scan row for ハイアルチ or HAIARUCHI
        delivery_address = ""
        for cell in row:
            if 'ハイアルチ' in cell or 'HAIARUCHI' in cell.upper():
                delivery_address = cell
                break

        if not order_id or not product_name:
            continue

        store = detect_store_from_address(delivery_address)
        short_name = _shorten_product_name(product_name)

        orders.append({
            "order_date": order_date,
            "order_id": order_id,
            "store_name": store,
            "product_name": product_name,
            "short_name": short_name,
            "amount": item_amount or order_total,
            "order_total": order_total,
            "payment_date": payment_date,
            "delivery_address": delivery_address,
        })

    return orders


def render(user=None):
    st.header("📤 データ取込")

    is_admin = user and user.get("role") == "admin"

    if is_admin:
        tab_payroll, tab_expense, tab_revenue, tab_budget = st.tabs(["💰 人件費", "🧾 経費", "📈 売上", "🎯 予算"])
    else:
        tab_expense, tab_revenue, tab_budget = st.tabs(["🧾 経費", "📈 売上", "🎯 予算"])
        tab_payroll = None

    # ─── 人件費 Upload ──────────────────────────────────────────
    if tab_payroll is not None:
      with tab_payroll:
        st.subheader("人件費データ取込")
        st.caption("クラウド給与から出力した支給控除一覧表（Excel / CSV）をアップロード")

        uploaded_payroll = st.file_uploader(
            "ファイルをアップロード",
            type=["xlsx", "xls", "csv"],
            key="payroll_upload",
        )

        if uploaded_payroll is not None:
            filename = uploaded_payroll.name
            detected_year, detected_month = _detect_year_month_from_filename(filename)

            st.info(f"📄 **{filename}**")

            col1, col2 = st.columns(2)
            with col1:
                payroll_year = st.number_input(
                    "対象年",
                    min_value=2020, max_value=2030,
                    value=detected_year or 2026,
                    key="payroll_year",
                )
            with col2:
                payroll_month = st.number_input(
                    "対象月",
                    min_value=1, max_value=12,
                    value=detected_month or 2,
                    key="payroll_month",
                )

            if detected_year and detected_month:
                st.success(f"ファイル名から **{detected_year}年{detected_month}月** を検出しました")

            # Step 1: Parse
            if st.button("▶ 人件費データを解析する", type="primary", key="btn_payroll_parse"):
                with st.spinner("解析中..."):
                    file_bytes = uploaded_payroll.read()

                    if filename.endswith(".csv"):
                        records, unresolved = parse_payroll_csv(file_bytes, payroll_year, payroll_month)
                    else:
                        records, unresolved = parse_payroll_excel(file_bytes, payroll_year, payroll_month)

                st.session_state["payroll_records"] = records
                st.session_state["payroll_unresolved"] = unresolved
                st.session_state["payroll_meta"] = {"year": payroll_year, "month": payroll_month}
                st.session_state["payroll_file_bytes"] = file_bytes
                st.session_state["payroll_filename"] = filename

            # Step 2: Show results and handle unresolved
            if "payroll_unresolved" in st.session_state:
                unresolved = st.session_state["payroll_unresolved"]
                records = st.session_state["payroll_records"]
                meta = st.session_state["payroll_meta"]

                if unresolved:
                    st.warning(f"⚠️ 店舗が不明な従業員が {len(unresolved)} 名います。店舗を選んでから「保存」を押してください。")

                    for emp in unresolved:
                        with st.container():
                            st.markdown(f"**{emp['employee_name']}** (ID: {emp['employee_id']}, {emp['contract_type']}, ¥{emp['gross_total']:,.0f})")
                            selected_store = st.selectbox(
                                f"店舗 — {emp['employee_name']}",
                                STORE_OPTIONS_WITH_HQ,
                                key=f"assign_{emp['employee_id']}",
                                label_visibility="collapsed",
                            )
                            emp["_selected_store"] = selected_store

                if records:
                    st.info(f"解析完了: {len(records)}件（振り分け済み）+ {len(unresolved)}件（未登録）")

                # Step 3: Save (with overwrite check)
                st.markdown("---")
                existing_count = check_payroll_exists(meta["year"], meta["month"])
                if existing_count > 0:
                    st.warning(f"⚠️ {meta['year']}年{meta['month']}月の人件費データが既に{existing_count}件あります。上書きしますか？")
                    ow_col1, ow_col2 = st.columns(2)
                    with ow_col1:
                        payroll_confirm = st.button("はい（上書き保存）", type="primary", key="btn_payroll_overwrite_yes")
                    with ow_col2:
                        if st.button("いいえ（キャンセル）", key="btn_payroll_overwrite_no"):
                            for key in ["payroll_records", "payroll_unresolved", "payroll_meta", "payroll_file_bytes"]:
                                if key in st.session_state:
                                    del st.session_state[key]
                            st.rerun()
                else:
                    payroll_confirm = st.button("💾 この内容で保存する", type="primary", key="btn_payroll_save_new")

                if payroll_confirm:
                    # Register unresolved employees first
                    for emp in unresolved:
                        store = emp.get("_selected_store", STORE_OPTIONS_WITH_HQ[0])
                        emp_id = int(emp["employee_id"])
                        upsert_override(emp_id, store, 100)

                    if unresolved and "payroll_file_bytes" in st.session_state:
                        # Re-parse with new registrations
                        file_bytes = st.session_state["payroll_file_bytes"]
                        fn = st.session_state.get("payroll_filename", "")
                        if fn.endswith(".csv"):
                            records, _ = parse_payroll_csv(file_bytes, meta["year"], meta["month"])
                        else:
                            records, _ = parse_payroll_excel(file_bytes, meta["year"], meta["month"])

                    save_payroll_data(records)
                    st.success(f"✅ {meta['year']}年{meta['month']}月の人件費データを保存しました（{len(records)}件）")

                    # Summary
                    df = pd.DataFrame(records)
                    summary = df.groupby("store_name").agg(
                        人数=("employee_id", "nunique"),
                        課税支給合計=("taxable_total", "sum"),
                        総勤務時間=("scheduled_hours", "sum"),
                    ).reset_index()
                    summary.columns = ["店舗", "人数", "課税支給合計", "総勤務時間"]
                    summary["課税支給合計"] = summary["課税支給合計"].apply(lambda x: f"¥{x:,.0f}")
                    summary["総勤務時間"] = summary["総勤務時間"].apply(lambda x: f"{x:,.0f}h")
                    st.dataframe(summary, use_container_width=True, hide_index=True)

                    # Clear session state
                    for key in ["payroll_records", "payroll_unresolved", "payroll_meta", "payroll_file_bytes"]:
                        if key in st.session_state:
                            del st.session_state[key]

    # ─── 経費 Upload ──────────────────────────────────────────
    with tab_expense:
        st.subheader("経費データ取込")
        st.caption("① Amazon注文履歴 → ② PayPay銀行CSV の順にアップロード")

        # Step 0: 店舗選択
        if is_admin:
            expense_store = st.selectbox("対象店舗", STORES, key="expense_store")
        else:
            expense_store = user.get("store_name", STORES[0])
            st.info(f"対象店舗: **{expense_store}**")

        st.markdown("---")

        # ── Step ①: Amazon注文履歴 ──────────────────────────
        st.markdown("### ① Amazon注文履歴（内訳データ）")

        amazon_just_imported = st.session_state.get("amazon_just_imported", False)

        skip_amazon = st.checkbox(
            "Amazonデータをスキップ（内訳不要の場合）",
            value=False,
            key="skip_amazon",
        )

        if not skip_amazon and not amazon_just_imported:
            uploaded_amazon = st.file_uploader(
                "Amazon注文履歴CSVをアップロード",
                type=["csv"],
                key="amazon_upload",
            )

            if uploaded_amazon is not None:
                if st.button("▶ 取り込む", type="primary", key="btn_amazon_import"):
                    with st.spinner("Amazon CSV 解析中..."):
                        amazon_bytes = uploaded_amazon.read()
                        amazon_records = parse_amazon_csv(amazon_bytes)

                    if amazon_records:
                        save_amazon_orders(amazon_records)
                        st.session_state["amazon_just_imported"] = True
                        st.session_state["amazon_import_msg"] = f"✅ Amazon注文 {len(amazon_records)} 件を取り込みました"
                        st.rerun()
                    else:
                        st.warning("取り込めるデータが見つかりませんでした。")

        amazon_ready = skip_amazon or amazon_just_imported

        if amazon_just_imported:
            if "amazon_import_msg" in st.session_state:
                st.success(st.session_state["amazon_import_msg"])
                del st.session_state["amazon_import_msg"]
            else:
                st.success("✅ Amazon注文データ取込完了")

        if not amazon_ready:
            st.info("Amazon注文履歴CSVをアップロードするか、スキップにチェックを入れてください。")

        # ── Step ②: PayPay銀行CSV ──────────────────────────
        st.markdown("---")
        st.markdown("### ② PayPay銀行CSV")

        if not amazon_ready:
            st.warning("先に①のAmazonデータを取り込むか、スキップしてください。")
        else:
            uploaded_expense = st.file_uploader(
                "PayPay銀行 入出金明細CSVをアップロード",
                type=["csv"],
                key="expense_upload",
            )

            if uploaded_expense is not None:
                file_bytes = uploaded_expense.read()
                if file_bytes:
                    st.session_state["expense_file_bytes"] = file_bytes
                    st.session_state["expense_file_name"] = uploaded_expense.name

            if "expense_file_bytes" in st.session_state:
                st.info(f"📄 **{st.session_state.get('expense_file_name', '')}** → **{expense_store}**")

                if st.button("▶ 解析する", type="primary", key="btn_expense_parse"):
                    records = None
                    parse_error = None
                    with st.spinner("解析中..."):
                        file_bytes = st.session_state["expense_file_bytes"]
                        for enc in ["cp932", "utf-8", "utf-8-sig"]:
                            try:
                                records = parse_expense_csv(file_bytes, enc)
                                break
                            except UnicodeDecodeError:
                                continue
                            except ValueError as e:
                                parse_error = str(e)
                                break

                    if parse_error:
                        st.error(parse_error)
                    elif records and len(records) > 0:
                        expense_year = records[0]["year"]
                        expense_month = records[0]["month"]

                        from database import find_breakdown_rule
                        for r in records:
                            r["store_name"] = expense_store
                            # First try breakdown rules (manual saves), then Amazon
                            rule = find_breakdown_rule(r["description"], int(r["amount"]))
                            if rule:
                                r["breakdown"] = rule
                            else:
                                r["breakdown"] = match_amazon_breakdown(
                                    r["description"], r["amount"], r["day"], r["month"], r["year"], expense_store
                                )

                        st.session_state["expense_records"] = records
                        st.session_state["expense_meta"] = {
                            "store": expense_store, "year": expense_year, "month": expense_month
                        }
                        del st.session_state["expense_file_bytes"]
                        if "expense_file_name" in st.session_state:
                            del st.session_state["expense_file_name"]
                    else:
                        st.warning("取り込めるデータが見つかりませんでした。")

            # Step 2: Show parsed results and allow classification
            if "expense_records" in st.session_state:
                records = st.session_state["expense_records"]
                meta = st.session_state["expense_meta"]

                classified = [r for r in records if r["category"] is not None]
                unclassified = [r for r in records if r["category"] is None]

                st.markdown(f"**解析結果:** {len(classified)}件分類済み / {len(unclassified)}件未分類")

                # Show classified summary
                if classified:
                    with st.expander(f"✅ 分類済み（{len(classified)}件）", expanded=False):
                        df_c = pd.DataFrame(classified)
                        exp_only = df_c[df_c["is_revenue"] == 0]
                        if not exp_only.empty:
                            summary = exp_only.groupby("category")["amount"].sum().sort_values(ascending=False).reset_index()
                            summary.columns = ["勘定科目", "合計金額"]
                            summary["合計金額"] = summary["合計金額"].apply(lambda x: f"¥{x:,.0f}")
                            st.dataframe(summary, use_container_width=True, hide_index=True)

                        # Show Amazon breakdown info
                        has_breakdown = [r for r in classified if r.get("breakdown")]
                        if has_breakdown:
                            st.markdown("**Amazon内訳付き:**")
                            for r in has_breakdown:
                                st.caption(f"  {r['day']}日 ¥{r['amount']:,.0f} → {r['breakdown']}")

                # Show unclassified items for manual selection
                if unclassified:
                    st.warning(f"⚠️ 以下の {len(unclassified)} 件の勘定科目を選んでください")
                    st.caption("選んだ勘定科目は自動分類ルールに保存され、次回から自動判定されます")
                    cat_options = ["（未選択）"] + EXPENSE_CATEGORIES

                    for i, rec in enumerate(unclassified):
                        uc_col1, uc_col2, uc_col3 = st.columns([4, 2, 3])
                        with uc_col1:
                            st.text(f"{rec['day']}日 — {rec['description']}")
                        with uc_col2:
                            if rec["amount"] > 0:
                                st.text(f"支払: ¥{rec['amount']:,.0f}")
                            else:
                                st.text(f"入金: ¥{rec['deposit']:,.0f}")
                        with uc_col3:
                            selected_cat = st.selectbox(
                                f"科目",
                                cat_options,
                                key=f"exp_cat_{i}",
                                label_visibility="collapsed",
                            )
                            if selected_cat != "（未選択）":
                                rec["category"] = selected_cat
                                # Auto-save to expense classification rules
                                upsert_expense_rule(rec["description"].strip(), selected_cat)

                # Save with overwrite confirmation
                st.markdown("---")
                existing_expense = check_expense_exists(meta["year"], meta["month"], meta["store"])
                need_confirm_exp = existing_expense > 0 and "expense_confirmed" not in st.session_state

                def _do_save_expense(records, meta):
                    save_expense_data(records)
                    saved_count = len([r for r in records if r["category"] is not None])
                    unsaved_count = len([r for r in records if r["category"] is None])
                    st.session_state["expense_save_msg"] = f"✅ **{meta['store']}** {meta['year']}年{meta['month']}月の経費データを保存しました（{saved_count}件分類済み、{unsaved_count}件未分類）"
                    for key in ["expense_records", "expense_meta", "expense_file_bytes",
                                "expense_filename", "amazon_just_imported", "expense_confirmed"]:
                        if key in st.session_state:
                            del st.session_state[key]
                    st.rerun()

                if need_confirm_exp:
                    st.warning(f"⚠️ **{meta['store']}** {meta['year']}年{meta['month']}月の経費データ（{existing_expense}件）が既に登録されています。上書きされますが、よろしいですか？")
                    col_yes, col_no, _ = st.columns([1, 1, 4])
                    with col_yes:
                        if st.button("はい（上書き）", key="btn_expense_yes"):
                            _do_save_expense(records, meta)
                    with col_no:
                        if st.button("いいえ（キャンセル）", key="btn_expense_no"):
                            for key in ["expense_records", "expense_meta", "expense_file_bytes", "expense_filename"]:
                                if key in st.session_state:
                                    del st.session_state[key]
                            st.info("キャンセルしました")
                            st.rerun()
                else:
                    if st.button("💾 この内容で保存する", type="primary", key="btn_expense_save"):
                        _do_save_expense(records, meta)

        # Show save success message
        if "expense_save_msg" in st.session_state:
            st.success(st.session_state["expense_save_msg"])
            del st.session_state["expense_save_msg"]

    # ─── 売上・会員 Upload (hacomono) ─────────────────────────
    with tab_revenue:
        st.subheader("hacomonoデータ取込")
        st.caption("hacomonoから出力したCSVをアップロード（会員リスト ML001 / 売上集計 PA002 / 売上明細 PL001）")

        sub_member, sub_sales_detail, sub_summary, sub_square = st.tabs(["👥 会員データ (ML001)", "🧾 売上明細 (PL001)", "📊 月次サマリ (MA002)", "💳 Square売上"])
        sub_sales = None  # PA002 deprecated

        # ─── 会員データ (ML001) ────────────────────────────
        with sub_member:
            st.markdown("#### 会員データ取込 (ML001)")
            st.caption("hacomono「メンバー一覧」CSVをアップロード — 常に最新データに上書きされます")

            from datetime import datetime as _dt
            _now = _dt.now()
            ml_year = _now.year
            ml_month = _now.month

            ml_store = st.selectbox("対象店舗", STORES, key="ml_store")

            uploaded_ml = st.file_uploader(
                "ML001 CSVをアップロード",
                type=["csv"],
                key="ml001_upload",
            )

            # Save file bytes to session_state immediately
            if uploaded_ml is not None:
                st.session_state["ml001_file_bytes"] = uploaded_ml.read()
                st.session_state["ml001_filename"] = uploaded_ml.name

            if "ml001_file_bytes" in st.session_state:
                st.info(f"📄 **{st.session_state.get('ml001_filename', '')}** → **{ml_store}**（最新データとして取込）")

            existing_ml = check_member_exists(ml_store)
            if existing_ml > 0:
                st.info(f"ℹ️ {ml_store} に既に{existing_ml}件の会員データがあります（上書きされます）")

            if st.button("▶ 会員データを取り込む", type="primary", key="btn_ml001"):
                if "ml001_file_bytes" in st.session_state:
                    file_bytes = st.session_state["ml001_file_bytes"]
                    try:
                        records, summary_info = _parse_ml001_csv(
                            file_bytes, ml_year, ml_month, ml_store
                        )

                        if records:
                            save_member_data(records)
                            st.success(
                                f"✅ **{ml_store}** の会員データを取り込みました"
                                f"（{summary_info['total']}名）"
                            )
                            if summary_info.get("empty_store_count", 0) > 0:
                                st.info(
                                    f"ℹ️ 店舗名が空の会員 {summary_info['empty_store_count']} 名も含めて取り込みました"
                                )

                            # Display summary
                            _render_ml001_summary(summary_info)
                            # Clear file bytes after successful save
                            for key in ["ml001_file_bytes", "ml001_filename"]:
                                if key in st.session_state:
                                    del st.session_state[key]
                        else:
                            st.warning("データが見つかりませんでした。ファイルの形式を確認してください。")

                    except Exception as e:
                        st.error(f"ファイルの読み込みに失敗しました: {e}")
                        import traceback
                        st.code(traceback.format_exc())
                else:
                    st.warning("CSVファイルをアップロードしてください。")

        # ─── 売上データ (PA002) — deprecated ───────────────────
        if False:  # PA002 tab disabled
          with sub_sales:
            st.markdown("#### 売上データ取込 (PA002)")
            st.caption("hacomono「売上集計」クエリ PA002 の CSV をアップロード（店舗ごとにアップロード）")

            col_store_r, col_year_r, col_month_r = st.columns(3)
            with col_store_r:
                rev_store = st.selectbox("対象店舗", STORES, key="rev_store")
            with col_year_r:
                rev_year = st.number_input("対象年", min_value=2020, max_value=2030, value=2026, key="rev_year")
            with col_month_r:
                rev_month = st.number_input("対象月", min_value=1, max_value=12, value=2, key="rev_month")

            uploaded_pa = st.file_uploader(
                "PA002 CSVをアップロード",
                type=["csv"],
                key="pa002_upload",
            )

            if uploaded_pa is not None:
                st.info(f"📄 **{uploaded_pa.name}** → **{rev_store}** / {rev_year}年{rev_month}月")

            if st.button("▶ 売上データを取り込む", type="primary", key="btn_pa002"):
                if uploaded_pa is not None:
                    file_bytes = uploaded_pa.read()
                    try:
                        # Try encodings
                        text = None
                        for enc in ["utf-8-sig", "utf-8", "cp932"]:
                            try:
                                text = file_bytes.decode(enc)
                                break
                            except UnicodeDecodeError:
                                continue
                        if text is None:
                            raise ValueError("CSVのエンコーディングを判定できませんでした")

                        reader = csv.reader(io.StringIO(text))
                        header = next(reader)

                        # Find column indices by header name
                        col_map = {}
                        for i, h in enumerate(header):
                            col_map[h.strip()] = i

                        data_row = next(reader, None)
                        if data_row is None:
                            st.warning("データ行が見つかりませんでした。")
                        else:
                            def _get_val(col_name):
                                idx = col_map.get(col_name)
                                if idx is not None and idx < len(data_row):
                                    try:
                                        return float(data_row[idx].strip().replace(",", ""))
                                    except (ValueError, TypeError):
                                        return 0.0
                                return 0.0

                            target_ym = data_row[col_map.get("対象年月", 0)].strip() if "対象年月" in col_map else ""
                            total_sales = _get_val("[総売上] 合計")
                            sales_amount = _get_val("[売上] 合計")
                            plan_sales_count = int(_get_val("[プラン売上] 件数"))
                            plan_sales_total = _get_val("[プラン売上] 合計")
                            plan_unit_price = _get_val("[プラン売上] 会員単価")

                            # Detect year/month from 対象年月 if it looks like "202602"
                            detected_year, detected_month = None, None
                            if len(target_ym) == 6:
                                try:
                                    detected_year = int(target_ym[:4])
                                    detected_month = int(target_ym[4:])
                                except ValueError:
                                    pass

                            if detected_year and detected_month:
                                st.info(f"CSVの対象年月: **{detected_year}年{detected_month}月**")

                            st.markdown("**検出した売上データ:**")
                            info_df = pd.DataFrame([{
                                "総売上合計": f"¥{total_sales:,.0f}",
                                "売上合計": f"¥{sales_amount:,.0f}",
                                "プラン売上件数": f"{plan_sales_count}件",
                                "プラン売上合計": f"¥{plan_sales_total:,.0f}",
                                "プラン会員単価": f"¥{plan_unit_price:,.0f}",
                            }])
                            st.dataframe(info_df, use_container_width=True, hide_index=True)

                            # Use detected year/month if available, otherwise use user-selected
                            save_year = detected_year or rev_year
                            save_month = detected_month or rev_month

                            rev_records = [{
                                "year": save_year,
                                "month": save_month,
                                "store_name": rev_store,
                                "category": "売上",
                                "amount": float(total_sales),
                                "member_count": plan_sales_count,
                                "note": f"PA002 | プラン売上: ¥{plan_sales_total:,.0f} | 会員単価: ¥{plan_unit_price:,.0f}",
                            }]
                            save_revenue_data(rev_records)
                            st.success(f"✅ **{rev_store}** {save_year}年{save_month}月の売上データを保存しました")

                    except Exception as e:
                        st.error(f"ファイルの読み込みに失敗しました: {e}")
                else:
                    st.warning("CSVファイルをアップロードしてください。")

        # ─── 売上明細 (PL001) ────────────────────────────
        with sub_sales_detail:
            st.markdown("#### 売上明細データ取込 (PL001)")
            st.caption("hacomono「売上明細」クエリ PL001 の CSV をアップロード（年月はCSVから自動検出）")

            if is_admin:
                pl_store = st.selectbox("対象店舗", STORES, key="pl_store")
            else:
                pl_store = user.get("store_name", STORES[0])
                st.info(f"対象店舗: **{pl_store}**")

            # Placeholder (will be overridden from CSV)
            pl_year = 2026
            pl_month = 1

            uploaded_pl = st.file_uploader(
                "PL001 CSVをアップロード",
                type=["csv"],
                key="pl001_upload",
            )

            if uploaded_pl is not None:
                st.info(f"📄 **{uploaded_pl.name}** → **{pl_store}**")

            if st.button("▶ 売上明細を取り込む", type="primary", key="btn_pl001"):
                if uploaded_pl is not None:
                    file_bytes = uploaded_pl.read()
                    try:
                        # Decode CSV
                        text = None
                        for enc in ["utf-8-sig", "utf-8", "cp932"]:
                            try:
                                text = file_bytes.decode(enc)
                                break
                            except UnicodeDecodeError:
                                continue
                        if text is None:
                            raise ValueError("CSVのエンコーディングを判定できませんでした")

                        reader = csv.reader(io.StringIO(text))
                        header = next(reader)

                        # Build column index map
                        hmap = {}
                        for i, h in enumerate(header):
                            hmap[h.strip()] = i

                        def _pl_get(row, col_name, default=""):
                            idx = hmap.get(col_name)
                            if idx is not None and idx < len(row):
                                return row[idx].strip()
                            return default

                        def _pl_int(row, col_name):
                            val = _pl_get(row, col_name, "0")
                            try:
                                return int(val.replace(",", ""))
                            except (ValueError, TypeError):
                                return 0

                        records = []
                        detected_year_pl = None
                        detected_month_pl = None

                        for row in reader:
                            if len(row) < 5:
                                continue

                            sale_id = _pl_get(row, "売上ID")
                            sale_date = _pl_get(row, "精算日時")
                            store_full = _pl_get(row, "購入店舗")
                            payment_method = _pl_get(row, "支払方法")
                            description = _pl_get(row, "摘要")
                            amount = _pl_int(row, "合計金額")
                            tax = _pl_int(row, "内税")
                            discount = _pl_int(row, "割引金額")

                            # Map store name
                            store_short = _map_hacomono_store(store_full) if store_full else pl_store

                            # Auto-detect year/month from first row's date
                            if detected_year_pl is None and sale_date:
                                dt = _parse_date_loose(sale_date)
                                if dt:
                                    detected_year_pl = dt.year
                                    detected_month_pl = dt.month

                            category = classify_sale_category(description, amount)

                            records.append({
                                "year": pl_year,  # will be overridden below if detected
                                "month": pl_month,
                                "store_name": store_short,
                                "sale_id": sale_id,
                                "sale_date": sale_date,
                                "payment_method": payment_method,
                                "description": description,
                                "category": category,
                                "amount": amount,
                                "tax": tax,
                                "discount": discount,
                            })

                        # Override year/month if detected
                        if detected_year_pl and detected_month_pl:
                            st.info(f"CSVから **{detected_year_pl}年{detected_month_pl}月** のデータを検出しました")
                            for r in records:
                                r["year"] = detected_year_pl
                                r["month"] = detected_month_pl
                            save_year_pl = detected_year_pl
                            save_month_pl = detected_month_pl
                        else:
                            save_year_pl = pl_year
                            save_month_pl = pl_month

                        if records:
                            save_sales_detail(records)

                            # Summary
                            total_amount = sum(r["amount"] for r in records)
                            cat_breakdown = {}
                            for r in records:
                                cat = r["category"]
                                cat_breakdown[cat] = cat_breakdown.get(cat, 0) + r["amount"]

                            st.success(
                                f"✅ **{pl_store}** {save_year_pl}年{save_month_pl}月の売上明細を取り込みました"
                                f"（{len(records)}件）"
                            )

                            st.markdown("---")
                            st.markdown("### 取込結果サマリー")

                            sk1, sk2 = st.columns(2)
                            with sk1:
                                st.metric("売上合計", f"¥{total_amount:,.0f}")
                            with sk2:
                                st.metric("取引件数", f"{len(records)}件")

                            # Category breakdown
                            st.markdown("**カテゴリ別内訳**")
                            cat_data = sorted(cat_breakdown.items(), key=lambda x: -x[1])
                            cat_df = pd.DataFrame(cat_data, columns=["カテゴリ", "金額"])
                            cat_df["構成比"] = cat_df["金額"].apply(
                                lambda x: f"{x / total_amount * 100:.1f}%" if total_amount != 0 else "0%"
                            )
                            cat_df["件数"] = [
                                sum(1 for r in records if r["category"] == cat) for cat, _ in cat_data
                            ]
                            cat_df["金額"] = cat_df["金額"].apply(lambda x: f"¥{x:,.0f}")
                            st.dataframe(cat_df, use_container_width=True, hide_index=True)
                        else:
                            st.warning("データが見つかりませんでした。ファイルの形式を確認してください。")

                    except Exception as e:
                        st.error(f"ファイルの読み込みに失敗しました: {e}")
                        import traceback
                        st.code(traceback.format_exc())
                else:
                    st.warning("CSVファイルをアップロードしてください。")

        # ─── 月次サマリ (MA002) ────────────────────────────
        with sub_summary:
            st.markdown("#### 月次サマリ取込 (MA002)")
            st.caption("hacomono「月次サマリ」クエリ MA002 の CSV をアップロード（年月はCSVから自動検出、複数月まとめてOK）")

            if is_admin:
                ma_store = st.selectbox("対象店舗", STORES, key="ma_store")
            else:
                ma_store = user.get("store_name", STORES[0])
                st.info(f"対象店舗: **{ma_store}**")

            uploaded_ma = st.file_uploader(
                "MA002 CSVをアップロード",
                type=["csv"],
                key="ma002_upload",
            )

            if uploaded_ma is not None:
                st.info(f"📄 **{uploaded_ma.name}** → **{ma_store}**")

            if st.button("▶ 月次サマリを取り込む", type="primary", key="btn_ma002"):
                if uploaded_ma is not None:
                    file_bytes = uploaded_ma.read()
                    try:
                        # Decode CSV
                        text = None
                        for enc in ["utf-8-sig", "utf-8", "cp932"]:
                            try:
                                text = file_bytes.decode(enc)
                                break
                            except UnicodeDecodeError:
                                continue
                        if text is None:
                            raise ValueError("CSVのエンコーディングを判定できませんでした")

                        reader = csv.reader(io.StringIO(text))
                        header = next(reader)

                        # Build column index map
                        hmap = {}
                        for i, h in enumerate(header):
                            hmap[h.strip()] = i

                        all_rows = list(reader)
                        if not all_rows:
                            st.warning("データ行が見つかりませんでした。")
                        else:
                            def _row_get(row, col_name, default=""):
                                idx = hmap.get(col_name)
                                if idx is not None and idx < len(row):
                                    return row[idx].strip()
                                return default

                            def _row_int(row, col_name):
                                val = _row_get(row, col_name, "0")
                                try:
                                    return int(val.replace(",", ""))
                                except (ValueError, TypeError):
                                    return 0

                            records_to_save = []
                            for data_row in all_rows:
                                target_ym = _row_get(data_row, "対象年月")
                                if len(target_ym) != 6:
                                    continue
                                try:
                                    save_year_ma = int(target_ym[:4])
                                    save_month_ma = int(target_ym[4:])
                                except ValueError:
                                    continue

                                records_to_save.append({
                                    "year": save_year_ma,
                                    "month": save_month_ma,
                                    "store_name": ma_store,
                                    "total_members": _row_int(data_row, "店舗在籍会員数"),
                                    "plan_subscribers": _row_int(data_row, "プラン契約者数"),
                                    "plan_subscribers_1st": _row_int(data_row, "プラン契約者数 (1日時点)"),
                                    "new_registrations": _row_int(data_row, "店舗在籍新規会員登録数"),
                                    "new_plan_applications": _row_int(data_row, "プラン新規申込数"),
                                    "new_plan_signups": _row_int(data_row, "プラン新規入会数"),
                                    "plan_changes": _row_int(data_row, "プラン変更数"),
                                    "suspensions": _row_int(data_row, "休会数"),
                                    "cancellations": _row_int(data_row, "退会数"),
                                    "cancellation_rate": _row_get(data_row, "退会率"),
                                })

                            if records_to_save:
                                for rec in records_to_save:
                                    save_monthly_summary([rec])

                                months_str = ", ".join(f"{r['year']}年{r['month']}月" for r in records_to_save)
                                st.success(f"✅ **{ma_store}** の月次サマリを取り込みました（{len(records_to_save)}ヶ月分: {months_str}）")

                                # Show summary for all months
                                st.markdown("---")
                                st.markdown("### 取込結果")
                                df_summary = pd.DataFrame([{
                                    "年月": f"{r['year']}年{r['month']:02d}月",
                                    "在籍会員数": r["total_members"],
                                    "プラン契約者数": r["plan_subscribers"],
                                    "新規入会": r["new_plan_signups"],
                                    "新規申込": r["new_plan_applications"],
                                    "退会": r["cancellations"],
                                    "退会率": r["cancellation_rate"],
                                } for r in records_to_save])
                                st.dataframe(df_summary, use_container_width=True, hide_index=True)
                            else:
                                st.warning("取り込めるデータが見つかりませんでした。")

                    except Exception as e:
                        st.error(f"ファイルの読み込みに失敗しました: {e}")
                        import traceback
                        st.code(traceback.format_exc())
                else:
                    st.warning("CSVファイルをアップロードしてください。")

        # ─── Square 売上サマリー ────────────────────────────
        with sub_square:
            st.markdown("#### Square売上サマリー取込")
            st.caption("Squareから出力した売上サマリーCSVをアップロード。ファイル名から期間を自動検出します（例: sales-summary-2026-02-01-2026-02-28.csv）")

            if is_admin:
                sq_store = st.selectbox("対象店舗", STORES, key="sq_store")
            else:
                sq_store = user.get("store_name", STORES[0])
                st.info(f"対象店舗: **{sq_store}**")

            uploaded_sq = st.file_uploader("Square CSVをアップロード", type=["csv"], key="square_upload")

            if uploaded_sq is not None:
                # Detect year/month from filename like sales-summary-2026-02-01-2026-02-28.csv
                fn = uploaded_sq.name
                m = re.search(r'(\d{4})-(\d{2})-\d{2}-\d{4}-\d{2}-\d{2}', fn)
                if m:
                    sq_year = int(m.group(1))
                    sq_month = int(m.group(2))
                    st.info(f"📄 **{fn}** → **{sq_store}** / {sq_year}年{sq_month}月")
                else:
                    st.error("ファイル名から年月を検出できませんでした。")
                    sq_year, sq_month = None, None

                if sq_year and st.button("▶ Square売上を取り込む", type="primary", key="btn_square"):
                    try:
                        file_bytes = uploaded_sq.read()
                        text = None
                        for enc in ["utf-8-sig", "utf-8", "cp932"]:
                            try:
                                text = file_bytes.decode(enc)
                                break
                            except UnicodeDecodeError:
                                continue

                        # Parse key-value CSV format
                        values = {}
                        for line in text.split("\n"):
                            parts = [p.strip().strip('"') for p in line.split(",")]
                            if len(parts) >= 2 and parts[0]:
                                values[parts[0].replace("\n", "").strip()] = parts[1]

                        def _yen(s):
                            if not s:
                                return 0
                            s = s.replace("¥", "").replace("￥", "").replace(",", "").replace("(", "-").replace(")", "").strip()
                            try:
                                return int(s)
                            except ValueError:
                                return 0

                        gross = _yen(values.get("総売上高", "0"))
                        net = _yen(values.get("合計（純額）", "0"))
                        fees = abs(_yen(values.get("手数料", "0")))
                        count = _yen(values.get("総売上数", "0"))

                        record = {
                            "store_name": sq_store,
                            "year": sq_year,
                            "month": sq_month,
                            "gross_sales": gross,
                            "net_sales": net,
                            "fees": fees,
                            "transaction_count": count,
                        }
                        save_square_sales([record])
                        st.success(f"✅ **{sq_store}** {sq_year}年{sq_month}月のSquare売上を取り込みました")
                        st.metric("総売上高", f"¥{gross:,}")
                        st.metric("取引件数", f"{count}件")
                    except Exception as e:
                        st.error(f"ファイルの読み込みに失敗しました: {e}")

    # ─── 予算 Upload ──────────────────────────────────────────
    with tab_budget:
        st.subheader("予算データ取込")
        st.caption("年度予算書（予算実績対比表のCSV形式）をアップロード。予算の列のみ読み取ります。")

        if is_admin:
            budget_store = st.selectbox("対象店舗", STORES, key="budget_store")
        else:
            budget_store = user.get("store_name", STORES[0])
            st.info(f"対象店舗: **{budget_store}**")

        col_fy, col_period = st.columns(2)
        with col_fy:
            fiscal_year = st.number_input("対象年度（決算年）", min_value=2020, max_value=2035, value=2026, key="budget_fy")
        with col_period:
            period_num = st.number_input("第○期", min_value=1, max_value=50, value=9, key="budget_period_num")
        st.caption(f"**{fiscal_year}年/第{period_num}期** = {fiscal_year-1}年10月〜{fiscal_year}年9月")

        uploaded_budget = st.file_uploader("予算CSVをアップロード", type=["csv"], key="budget_upload")

        if uploaded_budget is not None:
            if st.button("▶ 予算データを取り込む", type="primary", key="btn_budget"):
                try:
                    file_bytes = uploaded_budget.read()
                    text = None
                    for enc in ["utf-8-sig", "utf-8", "cp932"]:
                        try:
                            text = file_bytes.decode(enc)
                            break
                        except UnicodeDecodeError:
                            continue

                    reader = csv.reader(io.StringIO(text))
                    rows = list(reader)

                    # Fiscal year months: Oct(fy-1) - Sep(fy)
                    fy_months = [(fiscal_year - 1, m) for m in range(10, 13)] + [(fiscal_year, m) for m in range(1, 10)]

                    records = []
                    for row in rows:
                        if not row or not row[0].strip():
                            continue
                        item = row[0].strip()
                        if item not in BUDGET_ITEMS:
                            continue
                        # Each month has 4 columns: 予算/実績/予算差/予算比 starting from col index 1
                        for i, (y, m) in enumerate(fy_months):
                            col_idx = 1 + i * 4  # 予算 column
                            if col_idx < len(row):
                                val_str = row[col_idx].strip().replace(",", "").replace('"', '').replace(" ", "")
                                if val_str and val_str not in ("0", "-"):
                                    try:
                                        amount = int(val_str) * 1000
                                        records.append({
                                            "store_name": budget_store,
                                            "year": y,
                                            "month": m,
                                            "category": item,
                                            "amount": amount,
                                        })
                                    except ValueError:
                                        pass

                    if records:
                        existing = check_budget_exists(budget_store, fiscal_year)
                        if existing > 0:
                            st.warning(f"⚠️ {budget_store} の {fiscal_year}/9期 の予算データが既に登録されています（{existing}件）。上書きされます。")
                        saved = save_budget_data(records)
                        st.success(f"✅ **{budget_store}** {fiscal_year}/9期 の予算データを {saved}件 保存しました")

                        df_budget = pd.DataFrame(records)
                        summary = df_budget.groupby("category")["amount"].sum().sort_values(ascending=False).reset_index()
                        summary.columns = ["科目", "年間合計"]
                        summary["年間合計"] = summary["年間合計"].apply(lambda x: f"¥{x:,.0f}")
                        st.dataframe(summary, use_container_width=True, hide_index=True)
                    else:
                        st.warning("取り込めるデータが見つかりませんでした。CSVのフォーマットを確認してください。")
                except Exception as e:
                    st.error(f"ファイルの読み込みに失敗しました: {e}")
                    import traceback
                    st.code(traceback.format_exc())
